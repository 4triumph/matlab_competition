import math
from openpyxl import Workbook
# 定义位置信息的类
class Position:
    def __init__(self):
        self.x = 0.0
        self.w = 0.0
        self.depth = 0.0
        # self.fraction_coverage = 0.1
        self.d = 0.0

# 根据题目的得到的条件
alpha = 1.5
theta = 120
PI = 180
D = 110
original_x = D / math.tan(math.radians(alpha))
w1_angle = (PI - theta) / 2 - alpha
w2_angle = (PI - theta) / 2 + alpha
w3_angle = PI - w1_angle - alpha
w4_angle = PI - w2_angle
w5_angle = w2_angle - alpha
w6_angle = (PI - theta) / 2

# 计算海水深度
def cal_depth(x):
    return ((original_x - x) * D) / original_x

# # 计算左侧覆盖宽度
# def cal_w1(x):
#     return  x * math.sin(math.radians(theta / 2)) / math.sin(math.radians(w1_angle))
#
# # 计算右侧覆盖宽度
# def cal_w2(x):
#     return x * math.sin(math.radians(theta / 2)) / math.sin(math.radians(w2_angle))
def cal_w(x):
    return x * math.sin(math.radians(theta / 2)) / math.sin(math.radians(w6_angle))

def solve():
    positions = []  # 初始化一个空的列表
    x = -1852.0  # 初始 x 值
    while x <= 1852.0:  # 、
        position = Position()  # 创建一个新的 Position 对象
        position.x = x
        position.depth = cal_depth(x)
        # position.w = cal_w1(position.depth) + cal_w2(position.depth)
        position.w =
        position.d = (1 - 0.1) * position.w * math.sin(math.radians(w4_angle)) / math.sin(math.radians(w6_angle))  # 添加d属性到每个Position对象
        positions.append(position)  # 将新对象添加到列表中
        x += position.d  # 增加 x 值

    print(len(positions))
    for i in range(len(positions) - 1):
        print(f"该测线的覆盖宽度：{positions[i].w}")
        print(f"该一测线位置：{positions[i].x}")
        print(f"与后一条测线的距离：{positions[i].d}")

    workbook = Workbook()
    sheet: object = workbook.active

    # 将数据写入工作表
    for i, position in enumerate(positions, start=2):  # 从第二行开始写入数据
        sheet.cell(row=i, column=1, value=position.x)
        sheet.cell(row=i, column=2, value=position.depth)
        sheet.cell(row=i, column=3, value=position.w)
        sheet.cell(row=i, column=4, value=position.d)

    # 保存 Excel 文件
    workbook.save('output_data.xlsx')

if __name__ == "__main__":
    solve()
